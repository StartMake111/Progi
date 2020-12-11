import openpyxl
import datetime
import json
import os.path
from flask import Flask, request
app = Flask(__name__)
columnN,columnID,columnTime,columnItem,columnPrice = 1,2,3,4,5
if not(os.path.exists('data.xlsx')):    
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(1,1).value = 'N'
    sheet.cell(1,2).value = 'User ID'
    sheet.cell(1,3).value = 'Datetime'
    sheet.cell(1,4).value = 'Item'
    sheet.cell(1,5).value = 'Price'
    sheet.cell(25,25).value = 1
    sheet.cell(25,26).value = 1
    book.save('data.xlsx')
    book.close
@app.route('/', methods = ['POST', 'GET'])
def index():
    book = openpyxl.open('data.xlsx', read_only=True)
    sheet = book.active
    lastadd = sheet["Y25"].value
    nest = sheet["Z25"].value
    book.close
    timereq = datetime.datetime.now().time()
    if request.method == 'POST':
        userid = request.json["user_id"]
        cart = request.json["cart"]
        book = openpyxl.load_workbook('data.xlsx')
        sheet = book.active
        lastadd+=1
        sheet.cell(lastadd,columnN).value = nest
        sheet.cell(lastadd,columnID).value = userid
        sheet.cell(lastadd,columnTime).value = timereq
        print(lastadd)
        for item in cart:
            itemname = item['item']
            itemprice = item['price']
            sheet.cell(lastadd,columnItem).value = itemname
            sheet.cell(lastadd,columnPrice).value = itemprice
            lastadd+=1
        sheet.cell(25,25).value = lastadd-1
        sheet.cell(25,26).value = nest+1
        book.save('data.xlsx')
        book.close
        return 'POST'
    if request.method == 'GET':
        return "GET"
if __name__ == "__main__":
    app.run()